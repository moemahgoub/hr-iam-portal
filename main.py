# main.py
# ============================================================
# HR Provisioning Portal (FastAPI + Entra ID + Microsoft Graph)
#
# Purpose:
# - Simple IAM-style portal to submit, approve, and execute
#   onboarding/offboarding workflows with role-based access.
#
# Roles (Entra App Roles -> ID token "roles" claim):
# - HR.Requester  : submit onboarding/offboarding requests
# - HR.Approver   : approve/reject requests
# - IT.Clearance  : execute offboarding tasks
#
# Workflow:
# - submitted -> approve/reject
#   - onboard  -> approved -> executed/failed (background creates user)
#   - offboard -> it_pending -> executed/failed (IT executes)
#
# Security controls:
# - Server-side RBAC on POST endpoints (no UI-only security)
# - UI gate for all pages (must have at least one portal role)
# - Data-level access control to prevent ID guessing
# - CSRF protection for POST forms
# - One-time form token to prevent double submit
# - Security audit logging for auth/RBAC failures
# ============================================================

from __future__ import annotations

import json
import secrets
from datetime import datetime, timedelta
from typing import List, Optional
from io import BytesIO
from datetime import datetime
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from dotenv import load_dotenv
from fastapi import BackgroundTasks, FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware

import msal

# ------------------------------------------------------------
# Configuration
# ------------------------------------------------------------
# NOTE: Keep config values in .env locally, and in App Service App Settings later.
load_dotenv()

from config import (  # noqa: E402
    CLIENT_ID,
    CLIENT_SECRET,
    DATABASE_PATH,
    GRAPH_SCOPES,
    REDIRECT_URI,
    SESSION_SECRET,
    TENANT_DOMAIN,
    TENANT_ID,
)

# ------------------------------------------------------------
# Local modules (must exist in your project)
# ------------------------------------------------------------
from database import (  # noqa: E402
    audit_log,
    ensure_schema,
    get_conn,
    init_db,
    next_employee_number,
    now_iso,
    security_audit_log,
)

from graph_client import (  # noqa: E402
    create_user,
    disable_user,
    find_user_by_employee_id,
    remove_user_from_all_groups,
    revoke_signin_sessions,
    user_exists_by_upn,
)

# ------------------------------------------------------------
# App setup
# ------------------------------------------------------------
app = FastAPI()
templates = Jinja2Templates(directory="templates")

# Session cookies (local dev)
# NOTE: When deploying to Azure App Service (HTTPS), change https_only=True
app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET,
    same_site="lax",
    https_only=False,  # local dev (http://localhost)
    max_age=1800,  # 30 minutes
    session_cookie="hr_portal_session",
)
@app.get("/healthz")
def healthz():
    return {"status": "ok"}

# Database init
init_db(db_path=DATABASE_PATH) if "db_path" in init_db.__code__.co_varnames else init_db()
ensure_schema()

# ------------------------------------------------------------
# Entra ID / MSAL
# ------------------------------------------------------------
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
SCOPES = GRAPH_SCOPES.split()  # sign-in scopes (app roles come from ID token claims)


def msal_app() -> msal.ConfidentialClientApplication:
    return msal.ConfidentialClientApplication(
        client_id=CLIENT_ID,
        client_credential=CLIENT_SECRET,
        authority=AUTHORITY,
    )


# ------------------------------------------------------------
# Helpers: claims, identity, roles
# ------------------------------------------------------------
def claims(request: Request) -> dict:
    """Read ID token claims stored in the session after login."""
    return request.session.get("id_token_claims") or {}


def current_user_upn(request: Request) -> str:
    """Signed-in user UPN/email (empty string if missing)."""
    c = claims(request)
    upn = c.get("preferred_username") or c.get("upn") or c.get("email") or ""
    return (upn or "").strip()


def current_user_display_name(request: Request) -> str:
    """Display name for UI/audit. Falls back to UPN local part."""
    c = claims(request)
    name = (c.get("name") or "").strip()
    if name:
        return name

    upn = current_user_upn(request)
    if upn:
        return upn.split("@")[0] if "@" in upn else upn
    return ""


def get_user_roles(request: Request) -> List[str]:
    """Entra app roles from ID token 'roles' claim."""
    return claims(request).get("roles") or []


def require_login_page(request: Request) -> Optional[RedirectResponse]:
    """
    UI behavior:
    - no claims -> redirect to /signin
    - claims exist but missing identity -> clear session and redirect to /signin
    """
    c = claims(request)
    if not c:
        return RedirectResponse("/signin", status_code=303)

    if not current_user_upn(request):
        request.session.clear()
        return RedirectResponse("/signin?reason=session", status_code=303)

    return None


def require_role(request: Request, allowed: List[str]) -> Optional[JSONResponse]:
    """
    POST/API role enforcement (server-side).

    Returns:
    - JSONResponse if blocked
    - None if allowed
    """
    c = claims(request)
    if not c:
        security_audit_log(
            user_upn=None,
            action="auth.required",
            details="Blocked request: unauthenticated.",
            meta={"path": str(request.url.path), "required_roles": allowed},
        )
        return JSONResponse({"error": "Not logged in"}, status_code=401)

    roles = get_user_roles(request) or []
    user_upn = current_user_upn(request) or None

    if not roles:
        security_audit_log(
            user_upn=user_upn,
            action="rbac.denied",
            details="Blocked request: token has no roles.",
            meta={"path": str(request.url.path), "required_roles": allowed},
        )
        return JSONResponse({"error": "Forbidden"}, status_code=403)

    if not any(r in roles for r in allowed):
        security_audit_log(
            user_upn=user_upn,
            action="rbac.denied",
            details="Blocked request: missing required role.",
            meta={
                "path": str(request.url.path),
                "required_roles": allowed,
                "user_roles": roles,
            },
        )
        return JSONResponse({"error": "Forbidden"}, status_code=403)

    return None


# ------------------------------------------------------------
# UI page access gate (must have at least one portal role)
# ------------------------------------------------------------
ALLOWED_PORTAL_ROLES = {"HR.Requester", "HR.Approver", "IT.Clearance", "HR.Auditor"}


def require_portal_access_page(request: Request) -> Optional[HTMLResponse]:
    """
    UI protection for GET pages:
    - Not logged in -> redirect to /signin
    - Logged in but no portal roles -> access denied page
    - Has at least one portal role -> allow
    """
    deny = require_login_page(request)
    if deny:
        return deny

    roles = set(get_user_roles(request) or [])
    if roles.isdisjoint(ALLOWED_PORTAL_ROLES):
        security_audit_log(
            user_upn=current_user_upn(request) or None,
            action="rbac.portal_denied",
            details="Signed in but no portal roles.",
            meta={"user_roles": list(roles)},
        )
        return templates.TemplateResponse(
            "access_denied.html",
            {
                "request": request,
                "title": "Access denied",
                "message": "You are signed in, but you do not have access to this portal.",
            },
            status_code=403,
        )

    return None

# ------------------------------------------------------------
# Template helpers
# ------------------------------------------------------------
def fmt_dt(s: str | None) -> str:
    """Format ISO string to DD/MM/YYYY HH:MM."""
    if not s:
        return ""
    try:
        s2 = str(s).replace("Z", "+00:00")
        dt = datetime.fromisoformat(s2)
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(s)


def get_csrf_token(request: Request) -> str:
    """CSRF token stored in session."""
    token = request.session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        request.session["csrf_token"] = token
    return token


def verify_csrf(request: Request, form_token: str | None) -> None:
    """Validate CSRF token from form against session."""
    session_token = request.session.get("csrf_token")
    if not session_token or not form_token or form_token != session_token:
        raise HTTPException(status_code=403, detail="CSRF validation failed")


templates.env.globals["me_upn"] = current_user_display_name
templates.env.globals["me_roles"] = get_user_roles
templates.env.globals["fmt_dt"] = fmt_dt
templates.env.globals["csrf_token"] = get_csrf_token

# ------------------------------------------------------------
# UPN helpers
# ------------------------------------------------------------
def _slug(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "").replace("-", "").replace("_", "")

def _upn_exists_anywhere(candidate: str) -> bool:
    # DB check
    conn = get_conn()
    try:
        existing = conn.execute(
            """
            SELECT id FROM requests
            WHERE type='onboard'
              AND json_extract(payload_json, '$.userPrincipalName') = ?
              AND status IN ('submitted','approved','executed')
            LIMIT 1
            """,
            (candidate,),
        ).fetchone()
    finally:
        conn.close()

    if existing:
        return True

    # Entra check
    return user_exists_by_upn(candidate)


def generate_unique_upn(first: str, last: str, tenant_domain: str) -> str:
    first_s = _slug(first)
    last_s = _slug(last)
    if not first_s or not last_s:
        raise RuntimeError("Invalid name for UPN generation")

    # Try expanding first-name letters: s+ahmed, sa+ahmed, saj+ahmed...
    for n in range(1, min(len(first_s), 10) + 1):
        base = f"{first_s[:n]}{last_s}"
        candidate = f"{base}@{tenant_domain}"
        if not _upn_exists_anywhere(candidate):
            return candidate

    # Fallback: numeric suffix
    base = f"{first_s[:1]}{last_s}"
    for i in range(2, 100):
        candidate = f"{base}{i}@{tenant_domain}"
        if not _upn_exists_anywhere(candidate):
            return candidate

    raise RuntimeError("Unable to generate a unique UPN (too many duplicates)")


def find_recent_same_name_request(req_id: int, display_name: str, minutes: int = 60) -> Optional[int]:
    """If same displayName submitted recently, require approver confirmation."""
    conn = get_conn()
    try:
        rows = conn.execute(
            """
            SELECT id, created_at
            FROM requests
            WHERE type='onboard'
              AND id != ?
              AND json_extract(payload_json, '$.displayName') = ?
              AND status IN ('submitted','approved','executed')
            ORDER BY id DESC
            LIMIT 20
            """,
            (req_id, display_name),
        ).fetchall()

        now = datetime.utcnow()
        for r in rows:
            try:
                s2 = str(r["created_at"]).replace("Z", "+00:00")
                created = datetime.fromisoformat(s2).replace(tzinfo=None)
                if (now - created) <= timedelta(minutes=minutes):
                    return int(r["id"])
            except Exception:
                continue

        return None
    finally:
        conn.close()


def generate_temp_password() -> str:
    """
    Demo-safe password generation (better than a hardcoded password).
    In real production, deliver securely to the user.
    """
    return secrets.token_urlsafe(12) + "!"




# ------------------------------------------------------------
# Auth UI
# ------------------------------------------------------------
@app.get("/signin", response_class=HTMLResponse)
def signin_page(request: Request, reason: str | None = None):
    if reason == "session":
        msg = "Your session expired. Please sign in again."
    elif reason == "denied":
        msg = "Sign-in was cancelled or blocked. Please try again."
    elif reason == "logout":
        msg = "You have signed out."
    else:
        msg = "Please sign in to continue."

    return templates.TemplateResponse(
        "signin.html",
        {"request": request, "title": "Sign in", "message": msg},
    )


@app.get("/login")
def login():
    url = msal_app().get_authorization_request_url(
        scopes=SCOPES,
        redirect_uri=REDIRECT_URI,
        prompt="select_account",
    )
    return RedirectResponse(url, status_code=303)


@app.get("/auth/callback")
def auth_callback(request: Request, code: str | None = None):
    if not code:
        request.session.clear()
        security_audit_log(
            user_upn=None,
            action="auth.failed",
            details="Missing authorization code in callback.",
            meta={"reason": "missing_code"},
        )
        return RedirectResponse("/signin?reason=denied", status_code=303)

    result = msal_app().acquire_token_by_authorization_code(
        code=code,
        scopes=SCOPES,
        redirect_uri=REDIRECT_URI,
    )

    if "id_token_claims" not in result:
        request.session.clear()
        security_audit_log(
            user_upn=None,
            action="auth.failed",
            details="Sign-in failed during token exchange.",
            meta={
                "error": result.get("error"),
                "suberror": result.get("suberror"),
                "error_codes": result.get("error_codes"),
            },
        )
        reason = "denied" if (result.get("error") or "").lower() == "access_denied" else "session"
        return RedirectResponse(f"/signin?reason={reason}", status_code=303)

    request.session["id_token_claims"] = result["id_token_claims"]
    request.session["last_login_at"] = now_iso()
    return RedirectResponse("/", status_code=303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/signin?reason=logout", status_code=303)


# ------------------------------------------------------------
# Background execution: ONBOARD (create user)
# ------------------------------------------------------------
def execute_onboard_request(req_id: int) -> None:
    """Runs after HR approves onboarding (idempotent: only runs if status == 'approved')."""
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT id, type, status, payload_json FROM requests WHERE id=?",
            (req_id,),
        ).fetchone()

        if not row or row["type"] != "onboard" or row["status"] != "approved":
            return

        payload = json.loads(row["payload_json"] or "{}")
        emp_no = payload.get("employee_number")
        upn = payload.get("userPrincipalName")

        audit_log(req_id, "system", "execute.start", f"Creating user (emp={emp_no}, upn={upn})")

        graph_payload = {
            "accountEnabled": True,
            "displayName": payload["displayName"],
            "mailNickname": upn.split("@")[0],
            "userPrincipalName": upn,
            "employeeId": str(emp_no),
            "passwordProfile": {
                "forceChangePasswordNextSignIn": True,
                "password": payload["temp_password"],
            },
            "department": payload.get("department"),
            "jobTitle": payload.get("jobTitle"),
        }

        result = create_user(graph_payload)

        conn.execute(
            "UPDATE requests SET status=?, executed_at=?, result_json=? WHERE id=?",
            ("executed", now_iso(), json.dumps(result), req_id),
        )
        conn.commit()

        audit_log(req_id, "system", "execute.success", f"User created: {result.get('userPrincipalName', upn)}")

    except Exception as e:
        try:
            conn.execute(
                "UPDATE requests SET status=?, executed_at=?, result_json=? WHERE id=?",
                ("failed", now_iso(), json.dumps({"error": str(e)}), req_id),
            )
            conn.commit()
        except Exception:
            pass

        audit_log(req_id, "system", "execute.failed", str(e))
    finally:
        conn.close()


# ------------------------------------------------------------
# Home
# ------------------------------------------------------------
@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    deny = require_portal_access_page(request)
    if deny:
        return deny

    roles = get_user_roles(request)
    pending_approvals = None
    it_queue_count = None

    conn = get_conn()
    try:
        if "HR.Approver" in roles:
            pending_approvals = conn.execute(
                "SELECT COUNT(*) AS c FROM requests WHERE status='submitted'"
            ).fetchone()["c"]

        if "IT.Clearance" in roles:
            it_queue_count = conn.execute(
                "SELECT COUNT(*) AS c FROM requests WHERE type='offboard' AND status='it_pending'"
            ).fetchone()["c"]
    finally:
        conn.close()

    return templates.TemplateResponse(
        "home.html",
        {
            "request": request,
            "title": "Home",
            "roles": roles,
            "pending_approvals": pending_approvals,
            "it_queue_count": it_queue_count,
        },
    )


# ============================================================
# UI ROUTES
# ============================================================
@app.get("/requests/new", response_class=HTMLResponse)
def new_request_page(request: Request):
    deny = require_portal_access_page(request)
    if deny:
        return deny

    roles = get_user_roles(request)
    if "HR.Requester" not in roles:
        return templates.TemplateResponse(
            "access_denied.html",
            {"request": request, "title": "Access denied", "message": "You do not have access to create requests."},
            status_code=403,
        )

    draft_emp_no = next_employee_number()

    # One-time token to stop refresh/double submit (onboard form)
    form_token = secrets.token_urlsafe(24)
    request.session["onboard_form_token"] = form_token

    return templates.TemplateResponse(
        "new_request.html",
        {
            "request": request,
            "title": "New Request",
            "draft_emp_no": draft_emp_no,
            "form_token": form_token,
        },
    )


@app.get("/requests", response_class=HTMLResponse)
def list_requests(request: Request, status: str | None = None):
    """
    Least-privilege list view:
    - HR.Requester: only own requests
    - HR.Approver : submitted + ones they approved
    - IT.Clearance: offboard items in IT queue
    """
    deny = require_portal_access_page(request)
    if deny:
        return deny

    roles = get_user_roles(request)
    user_upn = current_user_upn(request)

    params: list = []
    where: list[str] = []

    if "HR.Approver" in roles:
        where.append("(status='submitted' OR approved_by=?)")
        params.append(current_user_display_name(request))
    elif "IT.Clearance" in roles:
        where.append("type='offboard'")
        where.append("status='it_pending'")
    elif "HR.Requester" in roles:
        where.append("created_by=?")
        params.append(user_upn)

    if status:
        where.append("status=?")
        params.append(status)

    where_sql = (" WHERE " + " AND ".join(where)) if where else ""

    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT id,type,status,created_by,created_at,approved_by,approved_at,executed_at "
            f"FROM requests{where_sql} ORDER BY id DESC",
            params,
        ).fetchall()
    finally:
        conn.close()

    return templates.TemplateResponse(
        "requests_list.html",
        {"request": request, "rows": rows, "title": "Requests", "status": status or ""},
    )


@app.get("/it/requests", response_class=HTMLResponse)
def it_queue(request: Request):
    deny = require_portal_access_page(request)
    if deny:
        return deny

    if "IT.Clearance" not in get_user_roles(request):
        return templates.TemplateResponse(
            "access_denied.html",
            {
                "request": request,
                "title": "Access denied",
                "message": "You do not have access to the IT clearance queue.",
            },
            status_code=403,
        )

    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT id,type,status,created_by,created_at,approved_by,approved_at,executed_at "
            "FROM requests WHERE type='offboard' AND status='it_pending' ORDER BY id DESC"
        ).fetchall()
    finally:
        conn.close()

    return templates.TemplateResponse(
        "requests_list.html",
        {"request": request, "title": "IT Clearance Queue", "rows": rows, "status": "it_pending"},
    )



def _friendly_actor_name(v: str | None) -> str:
    raw = (v or "").split("@")[0]
    if not raw:
        return "-"
    if raw.startswith("hr.requester"):
        suffix = raw.replace("hr.requester", "")
        return f"HR Requester {suffix}".strip()
    if raw.startswith("hr.approver"):
        suffix = raw.replace("hr.approver", "")
        return f"HR Approver {suffix}".strip()
    if raw.startswith("it.clearance"):
        suffix = raw.replace("it.clearance", "")
        return f"IT Clearance {suffix}".strip()
    if raw == "system":
        return "System"
    return raw.replace(".", " ").replace("_", " ").title()


def _friendly_action(a: str | None) -> str:
    m = {
        "submit": "Request submitted",
        "approve": "Approved",
        "reject": "Rejected",
        "bg.scheduled": "Execution scheduled",
        "execute.start": "Creating user account",
        "execute.success": "User account created",
        "execute.failed": "User creation failed",
        "it.execute.start": "IT started offboarding",
        "it.disable": "Account disabled",
        "it.revoke": "Sessions revoked",
        "it.groups.remove": "Groups removed",
        "it.execute.success": "Offboarding completed",
        "it.execute.failed": "Offboarding failed",
        "upn.adjust": "UPN adjusted",
    }
    return m.get((a or "").strip(), (a or "").strip())


def _fmt_excel_dt(s: str | None) -> str:
    if not s:
        return ""
    try:
        s2 = str(s).replace("Z", "+00:00")
        dt = datetime.fromisoformat(s2)
        return dt.strftime("%d/%m/%Y %H:%M:%S")
    except Exception:
        return str(s)


def _safe_json_load(s: str | None) -> dict:
    if not s:
        return {}
    try:
        return json.loads(s)
    except Exception:
        return {}


def _split_name(full_name: str | None) -> tuple[str, str]:
    full = (full_name or "").strip()
    if not full:
        return "", ""
    parts = full.split()
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


# ------------------------------------------------------------
# Audit Export (Excel)
# Access policy:
# - Must be signed in and have portal access
# - Must have HR.Auditor role (strict)
# - Returns enterprise-friendly .xlsx export
# ------------------------------------------------------------
@app.get("/audit/export")
def export_audit_excel(request: Request):
    """
    Enterprise-friendly Excel export (.xlsx)
    Access: HR.Auditor only
    """
    # 1) Login + portal access gate (redirects/HTML deny page)
    deny = require_portal_access_page(request)
    if deny:
        return deny

    # 2) Strict role check for browser route (HTML deny page, not JSON)
    roles = get_user_roles(request) or []
    if "HR.Auditor" not in roles:
        security_audit_log(
            user_upn=current_user_upn(request) or None,
            action="rbac.audit_export_denied",
            details="Blocked audit export: missing HR.Auditor role.",
            meta={
                "path": str(request.url.path),
                "user_roles": roles,
            },
        )
        return templates.TemplateResponse(
            "access_denied.html",
            {
                "request": request,
                "title": "Access denied",
                "message": "You do not have access to export audit data.",
            },
            status_code=403,
        )

    # 3) Read requests + latest audit event per request
    conn = get_conn()
    try:
        req_rows = conn.execute(
            """
            SELECT
                r.id,
                r.type,
                r.status,
                r.payload_json,
                r.result_json,
                r.created_by,
                r.created_at,
                r.approved_by,
                r.approved_at,
                r.executed_at
            FROM requests r
            ORDER BY r.id DESC
            """
        ).fetchall()

        # Latest audit row for each request (based on max audit.id)
        last_audit_map = {}
        audit_rows = conn.execute(
            """
            SELECT a.request_id, a.actor, a.action, a.details, a.at
            FROM audit a
            INNER JOIN (
                SELECT request_id, MAX(id) AS max_id
                FROM audit
                GROUP BY request_id
            ) x
                ON a.request_id = x.request_id
               AND a.id = x.max_id
            """
        ).fetchall()

        for a in audit_rows:
            last_audit_map[a["request_id"]] = a

    finally:
        conn.close()

    # 4) Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Export"

    headers = [
        "Request ID",
        "Request Type",
        "Status",
        "Employee ID",
        "First Name",
        "Last Name",
        "Full Name",
        "Email / UPN",
        "Department",
        "Job Title",
        "Remove Groups",
        "Created By",
        "Created At",
        "Approved / Rejected By",
        "Approved / Rejected At",
        "Executed At",
        "Last Action",
        "Last Action By",
        "Last Action Time",
        "Result / Error",
    ]
    ws.append(headers)

    # Header style (enterprise readable)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 5) Populate rows
    for r in req_rows:
        payload = _safe_json_load(r["payload_json"])
        result = _safe_json_load(r["result_json"])

        emp_id = payload.get("employee_number", "")
        full_name = payload.get("displayName", "") or ""
        first_name, last_name = _split_name(full_name)

        upn = payload.get("userPrincipalName", "") or ""
        dept = payload.get("department", "") or payload.get("Department", "") or ""
        job = payload.get("jobTitle", "") or payload.get("job_title", "") or ""

        remove_groups = ""
        if r["type"] == "offboard":
            remove_groups = "Yes" if bool(payload.get("remove_groups", False)) else "No"

        # Friendly actor names
        created_by = _friendly_actor_name(r["created_by"])
        approved_by = _friendly_actor_name(r["approved_by"])

        # Latest audit info
        last_audit = last_audit_map.get(r["id"])
        last_action = _friendly_action(last_audit["action"]) if last_audit else ""
        last_actor = _friendly_actor_name(last_audit["actor"]) if last_audit else ""
        last_action_time = _fmt_excel_dt(last_audit["at"]) if last_audit else ""

        # Friendly result/error summary
        result_text = ""
        if r["status"] == "failed":
            result_text = result.get("error", "") if isinstance(result, dict) else str(result or "")
        elif r["status"] == "rejected":
            result_text = result.get("reason", "") if isinstance(result, dict) else str(result or "")
        elif r["status"] == "executed":
            if r["type"] == "onboard":
                result_text = "User created successfully"
            else:
                groups_removed = result.get("groups_removed", 0) if isinstance(result, dict) else 0
                result_text = f"Offboarding completed (groups removed: {groups_removed})"

        ws.append([
            r["id"],
            str(r["type"]).title(),
            str(r["status"]).replace("_", " ").title(),
            emp_id,
            first_name,
            last_name,
            full_name,
            upn,
            dept,
            job,
            remove_groups,
            created_by,
            _fmt_excel_dt(r["created_at"]),
            approved_by,
            _fmt_excel_dt(r["approved_at"]),
            _fmt_excel_dt(r["executed_at"]),
            last_action,
            last_actor,
            last_action_time,
            result_text,
        ])

    # 6) Sheet usability formatting
    ws.freeze_panes = "A2"                # Keep header visible
    ws.auto_filter.ref = ws.dimensions    # Filter on all columns

    widths = {
        "A": 10, "B": 14, "C": 16, "D": 12,
        "E": 16, "F": 18, "G": 24, "H": 34,
        "I": 18, "J": 22, "K": 14, "L": 20,
        "M": 20, "N": 22, "O": 22, "P": 22,
        "Q": 24, "R": 20, "S": 22, "T": 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Wrap long cells + top-align for readability
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 7) Return file as download
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"audit_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
@app.get("/requests/{req_id}", response_class=HTMLResponse)
def request_detail(req_id: int, request: Request):
    """
    Request Details page (secured):
    - Prevents ID guessing
    - Data-level access enforced based on role + request properties
    """
    deny = require_portal_access_page(request)
    if deny:
        return deny

    roles = get_user_roles(request)
    user_upn = current_user_upn(request)

    conn = get_conn()
    try:
        row = conn.execute("SELECT * FROM requests WHERE id=?", (req_id,)).fetchone()
        audit_rows = conn.execute(
            "SELECT actor, action, details, at FROM audit WHERE request_id=? ORDER BY id ASC",
            (req_id,),
        ).fetchall()
    finally:
        conn.close()

    if not row:
        return templates.TemplateResponse(
            "access_denied.html",
            {
                "request": request,
                "title": "Not found",
                "message": "This request does not exist, or you do not have access to it.",
            },
            status_code=404,
        )

    # Data-level access control
    allowed = False
    if "HR.Approver" in roles:
        allowed = True
    elif "IT.Clearance" in roles:
        allowed = (row["type"] == "offboard" and row["status"] in ("it_pending", "executed", "failed"))
    elif "HR.Requester" in roles:
        allowed = (row["created_by"] == user_upn)

    if not allowed:
        security_audit_log(
            user_upn=user_upn or None,
            action="rbac.data_denied",
            details="Blocked request detail view (data-level).",
            meta={"req_id": req_id, "user_roles": roles},
        )
        return templates.TemplateResponse(
            "access_denied.html",
            {
                "request": request,
                "title": "Access denied",
                "message": "You do not have access to view this request.",
            },
            status_code=403,
        )

    try:
        payload = json.loads(row["payload_json"] or "{}")
    except Exception:
        payload = {}

    fail_reason = ""
    if row["status"] == "failed" and row["result_json"]:
        try:
            rj = json.loads(row["result_json"])
            fail_reason = rj.get("error", "")
        except Exception:
            pass

    can_hr_approve_reject = (row["status"] == "submitted" and "HR.Approver" in roles)
    can_it_execute = (row["type"] == "offboard" and row["status"] == "it_pending" and "IT.Clearance" in roles)

    return templates.TemplateResponse(
        "request_detail.html",
        {
            "request": request,
            "title": f"Request #{req_id}",
            "row": row,
            "payload": payload,
            "audit_rows": audit_rows,
            "can_hr_approve_reject": can_hr_approve_reject,
            "can_it_execute": can_it_execute,
            "fail_reason": fail_reason,
        },
    )


# ============================================================
# SUBMIT REQUESTS (HR.Requester)
# ============================================================
@app.post("/requests/onboard")
async def submit_onboard(request: Request):
    deny = require_role(request, ["HR.Requester"])
    if deny:
        return deny

    form = await request.form()
    verify_csrf(request, form.get("csrf"))

    # One-time token to stop refresh/double submit
    token = (form.get("form_token") or "").strip()
    session_token = request.session.get("onboard_form_token")
    if not token or token != session_token:
        return HTMLResponse("This form was already submitted (or expired). Please open New Request again.", status_code=400)
    request.session["onboard_form_token"] = None

    first = (form.get("first_name") or "").strip()
    last = (form.get("last_name") or "").strip()
    dept = (form.get("department") or "").strip()
    job = (form.get("job_title") or "").strip()

    if not first or not last:
        return HTMLResponse("First and last name required", status_code=400)

    emp_no_raw = (form.get("employee_number") or "").strip()
    if not emp_no_raw.isdigit():
        return HTMLResponse("Invalid employee number", status_code=400)
    employee_number = int(emp_no_raw)

    upn = generate_unique_upn(first, last, TENANT_DOMAIN)

    payload = {
        "employee_number": employee_number,
        "displayName": f"{first} {last}",
        "department": dept,
        "jobTitle": job,
        "userPrincipalName": upn,
        "temp_password": generate_temp_password(),
    }

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO requests(type,status,payload_json,created_by,created_at) VALUES (?,?,?,?,?)",
            ("onboard", "submitted", json.dumps(payload), current_user_upn(request), now_iso()),
        )
        req_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()

    audit_log(req_id, current_user_display_name(request), "submit", f"Onboard submitted (emp={employee_number}, upn={upn})")
    return RedirectResponse(f"/requests/{req_id}", status_code=303)


@app.post("/requests/offboard")
async def submit_offboard(request: Request):
    deny = require_role(request, ["HR.Requester"])
    if deny:
        return deny

    form = await request.form()
    verify_csrf(request, form.get("csrf"))

    emp = (form.get("employee_number") or "").strip()
    remove_groups = (form.get("remove_groups") == "on")

    if not emp.isdigit():
        return HTMLResponse("Employee number must be numeric", status_code=400)

    existing = find_user_by_employee_id(int(emp))
    if not existing:
        return templates.TemplateResponse(
    "access_denied.html",
    {
        "request": request,
        "title": "Invalid employee number",
        "message": "Employee number not found.",
    },
    status_code=400,
)

    payload = {"employee_number": int(emp), "remove_groups": remove_groups}


    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO requests(type,status,payload_json,created_by,created_at) VALUES (?,?,?,?,?)",
            ("offboard", "submitted", json.dumps(payload), current_user_upn(request), now_iso()),
        )
        req_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()

    audit_log(
        req_id,
        current_user_display_name(request),
        "submit",
        f"Offboard submitted (emp={emp}, remove_groups={remove_groups})",
    )
    return RedirectResponse(f"/requests/{req_id}", status_code=303)


# ============================================================
# APPROVE / REJECT (HR.Approver)
# ============================================================
@app.post("/requests/{req_id}/approve")
async def approve(req_id: int, request: Request, background_tasks: BackgroundTasks):
    deny = require_role(request, ["HR.Approver"])
    if deny:
        return deny

    form = await request.form()
    verify_csrf(request, form.get("csrf"))

    confirm_same_name = (form.get("confirm_same_name") == "1")
    approver = current_user_display_name(request)

    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT id,type,status,payload_json FROM requests WHERE id=?",
            (req_id,),
        ).fetchone()

        if not row:
            return HTMLResponse("Not found", status_code=404)

        if row["status"] != "submitted":
            return HTMLResponse("Request is not in submitted state", status_code=400)

        # OFFBOARD -> send to IT queue
        if row["type"] == "offboard":
            conn.execute(
                "UPDATE requests SET status=?, approved_by=?, approved_at=? WHERE id=?",
                ("it_pending", approver, now_iso(), req_id),
            )
            conn.commit()
            audit_log(req_id, approver, "approve", "Sent to IT clearance queue")
            return RedirectResponse(f"/requests/{req_id}", status_code=303)

        # ONBOARD -> validate + possible confirm
        try:
            payload = json.loads(row["payload_json"] or "{}")
        except Exception:
            payload = {}

        emp_no = payload.get("employee_number")
        upn = payload.get("userPrincipalName")
        display_name = (payload.get("displayName") or "").strip()

        if not emp_no or not upn or not display_name:
            return HTMLResponse("Invalid onboarding payload.", status_code=400)

        # Confirm duplicate display name
        if not confirm_same_name:
            dup_req = find_recent_same_name_request(req_id, display_name, minutes=60)
            if dup_req:
                return templates.TemplateResponse(
                    "confirm_duplicate_name.html",
                    {
                        "request": request,
                        "title": "Confirm duplicate name",
                        "req_id": req_id,
                        "display_name": display_name,
                        "dup_req_id": dup_req,
                    },
                    status_code=200,
                )

        # Block if employeeId exists in Entra
        existing_emp = find_user_by_employee_id(emp_no)
        if existing_emp:
            return HTMLResponse(
                f"Cannot approve: employee already exists (employeeId={emp_no}).",
                status_code=400,
            )

        # If UPN exists -> auto-adjust
        if user_exists_by_upn(upn):
            parts = display_name.split()
            first = parts[0] if parts else "user"
            last = " ".join(parts[1:]) if len(parts) > 1 else "user"

            new_upn = generate_unique_upn(first, last, TENANT_DOMAIN)
            payload["userPrincipalName"] = new_upn

            conn.execute(
                "UPDATE requests SET payload_json=? WHERE id=?",
                (json.dumps(payload), req_id),
            )
            conn.commit()

            audit_log(req_id, approver, "upn.adjust", f"{upn} -> {new_upn}")

        # Approve + background create user
        conn.execute(
            "UPDATE requests SET status=?, approved_by=?, approved_at=? WHERE id=?",
            ("approved", approver, now_iso(), req_id),
        )
        conn.commit()

        audit_log(req_id, approver, "approve", "Onboarding approved")
        background_tasks.add_task(execute_onboard_request, req_id)
        return RedirectResponse(f"/requests/{req_id}", status_code=303)

    finally:
        conn.close()


@app.post("/requests/{req_id}/reject")
async def reject(req_id: int, request: Request):
    deny = require_role(request, ["HR.Approver"])
    if deny:
        return deny

    form = await request.form()
    verify_csrf(request, form.get("csrf"))

    reason = (form.get("reason") or "").strip()
    if len(reason) < 3:
        return HTMLResponse("Reject reason required", status_code=400)

    approver = current_user_display_name(request)

    conn = get_conn()
    try:
        conn.execute(
            "UPDATE requests SET status=?, approved_by=?, approved_at=?, result_json=? WHERE id=?",
            ("rejected", approver, now_iso(), json.dumps({"reason": reason}), req_id),
        )
        conn.commit()
    finally:
        conn.close()

    audit_log(req_id, approver, "reject", reason)
    return RedirectResponse(f"/requests/{req_id}", status_code=303)


# ============================================================
# IT EXECUTE (IT.Clearance) - OFFBOARD ONLY
# ============================================================
@app.post("/requests/{req_id}/it-execute")
async def it_execute(req_id: int, request: Request):
    deny = require_role(request, ["IT.Clearance"])
    if deny:
        return deny

    form = await request.form()
    verify_csrf(request, form.get("csrf"))

    actor = current_user_display_name(request)

    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT id, payload_json, status, type FROM requests WHERE id=?",
            (req_id,),
        ).fetchone()

        if not row or row["type"] != "offboard" or row["status"] != "it_pending":
            return HTMLResponse("Invalid request state", status_code=400)

        payload = json.loads(row["payload_json"] or "{}")
        emp_no = payload.get("employee_number")
        remove_groups = bool(payload.get("remove_groups", False))

        audit_log(req_id, actor, "it.execute.start", f"Offboarding started (emp={emp_no})")

        user = find_user_by_employee_id(emp_no)
        if not user:
            raise RuntimeError("User not found in Entra ID")

        user_id = user["id"]

        disable_user(user_id)
        audit_log(req_id, actor, "it.disable", "Account disabled")

        revoke_signin_sessions(user_id)
        audit_log(req_id, actor, "it.revoke", "Sign-in sessions revoked")

        removed = 0
        if remove_groups:
            removed = remove_user_from_all_groups(user_id)
            audit_log(req_id, actor, "it.groups.remove", f"Removed from {removed} groups")

        result = {"employee_number": emp_no, "groups_removed": removed}

        conn.execute(
            "UPDATE requests SET status=?, executed_at=?, result_json=? WHERE id=?",
            ("executed", now_iso(), json.dumps(result), req_id),
        )
        conn.commit()

        audit_log(req_id, actor, "it.execute.success", "Offboarding completed")

    except Exception as e:
        try:
            conn.execute(
                "UPDATE requests SET status=?, executed_at=?, result_json=? WHERE id=?",
                ("failed", now_iso(), json.dumps({"error": str(e)}), req_id),
            )
            conn.commit()
        except Exception:
            pass

        audit_log(req_id, actor, "it.execute.failed", str(e))

    finally:
        conn.close()

    return RedirectResponse(f"/requests/{req_id}", status_code=303)